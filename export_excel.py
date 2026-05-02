from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _fmt_date(dt) -> str:
    return dt.strftime("%d/%m/%Y") if dt else "N/A"


def generate_executive_report(db_session) -> bytes:
    from models import Tender

    tenders = (
        db_session.query(Tender)
        .filter(Tender.status != "Perdu")
        .order_by(Tender.deadline)
        .all()
    )

    rows = []
    for t in tenders:
        analysis = t.llm_analysis or {}
        rows.append(
            {
                "Date Limite": _fmt_date(t.deadline),
                "Titre du Marché": t.title or "Sans titre",
                "Type (Travaux/Maintenance)": analysis.get("type_marche", "Non analysé"),
                "Score DEF": t.relevance_score or analysis.get("score_pertinence", 0),
                "Concurrents identifiés": ", ".join(
                    analysis.get("marques_concurrentes_citees", [])
                ) or "Aucun",
                "Lien source": t.source or "",
                "Statut actuel": t.status or "À qualifier",
                "Risques / Pénalités": analysis.get("risques_penalites") or "Non renseigné",
                "Date de publication": _fmt_date(t.publication_date),
            }
        )

    df = pd.DataFrame(rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Opportunités DEF OI")

        wb = writer.book
        ws = writer.sheets["Opportunités DEF OI"]

        # Header styling
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Row colour by status
        status_colors = {
            "À qualifier": "FFF2CC",
            "En cours": "DDEEFF",
            "Soumis": "D9EAD3",
            "Gagné": "B7D7A8",
        }
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status_val = row[6].value  # column index 6 = "Statut actuel"
            color = status_colors.get(status_val, "FFFFFF")
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Column widths
        col_widths = [14, 45, 22, 12, 30, 40, 16, 35, 18]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 35
        ws.freeze_panes = "A2"

        # Metadata sheet
        meta = wb.create_sheet("Infos Rapport")
        meta["A1"] = "Rapport généré par"
        meta["B1"] = "DEF Océan Indien — Outil Veille Commerciale"
        meta["A2"] = "Date de génération"
        meta["B2"] = datetime.now().strftime("%d/%m/%Y à %H:%M")
        meta["A3"] = "Périmètre"
        meta["B3"] = "Départements 974 (La Réunion) et 976 (Mayotte)"
        meta["A4"] = "Marchés exclus"
        meta["B4"] = "Statut « Perdu »"
        meta["A5"] = "Total inclus"
        meta["B5"] = len(rows)
        for cell in ["A1", "A2", "A3", "A4", "A5"]:
            meta[cell].font = Font(bold=True)
        meta.column_dimensions["A"].width = 22
        meta.column_dimensions["B"].width = 50

    output.seek(0)
    return output.getvalue()
